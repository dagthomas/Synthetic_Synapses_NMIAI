import os
import logging
import pdfplumber
from PIL import Image
import google.genai as genai

from config import GEMINI_MODEL

log = logging.getLogger(__name__)


def build_file_tools(files_dir: str) -> dict:
    """Build file extraction tools. files_dir is the per-request temp directory."""

    def extract_file_content(filename: str) -> dict:
        """Extract text content from a PDF or image file attachment.

        Args:
            filename: The name of the attached file to extract text from. Supports PDF and image files (PNG, JPG, JPEG).

        Returns:
            The extracted text content, or an error message.
        """
        filepath = os.path.join(files_dir, filename)
        if not os.path.exists(filepath):
            return {"error": True, "message": f"File not found: {filename}"}

        lower = filename.lower()
        if lower.endswith(".pdf"):
            return _extract_pdf(filepath)
        elif lower.endswith((".png", ".jpg", ".jpeg", ".webp")):
            return _extract_image(filepath)
        elif lower.endswith((".csv", ".txt", ".tsv")):
            return _extract_text(filepath)
        elif lower.endswith((".xlsx", ".xls")):
            return _extract_excel(filepath)
        else:
            # Try reading as plain text
            return _extract_text(filepath)

    return {"extract_file_content": extract_file_content}


def _extract_pdf(filepath: str) -> dict:
    """Extract text from PDF. Falls back to Gemini vision for scanned PDFs."""
    try:
        text_parts = []
        with pdfplumber.open(filepath) as pdf:
            for page in pdf.pages:
                page_text = page.extract_text()
                if page_text:
                    text_parts.append(page_text)
        full_text = "\n\n".join(text_parts)
        if full_text.strip():
            return {"text": full_text}

        # Fallback: scanned PDF — render pages to images and use Gemini vision
        log.info("PDF has no text layer, falling back to Gemini vision")
        return _extract_pdf_with_vision(filepath)
    except Exception as e:
        return {"error": True, "message": f"Failed to extract PDF: {str(e)}"}


def _extract_pdf_with_vision(filepath: str) -> dict:
    """Render PDF pages to images and extract text with Gemini vision."""
    try:
        images = []
        with pdfplumber.open(filepath) as pdf:
            for page in pdf.pages:
                img = page.to_image(resolution=200).original
                images.append(img)

        client = genai.Client()
        texts = []
        for img in images:
            response = client.models.generate_content(
                model=GEMINI_MODEL,
                contents=["Extract ALL text from this image exactly as written. Preserve formatting, numbers, and special characters:", img],
            )
            texts.append(response.text)
        return {"text": "\n\n".join(texts)}
    except Exception as e:
        return {"error": True, "message": f"Vision extraction failed: {str(e)}"}


def _extract_text(filepath: str) -> dict:
    """Extract content from text/CSV files."""
    try:
        for encoding in ("utf-8", "latin-1", "cp1252"):
            try:
                with open(filepath, "r", encoding=encoding) as f:
                    text = f.read()
                return {"text": text}
            except UnicodeDecodeError:
                continue
        return {"error": True, "message": "Could not decode text file"}
    except Exception as e:
        return {"error": True, "message": f"Failed to read text file: {str(e)}"}


def _extract_excel(filepath: str) -> dict:
    """Extract content from Excel files."""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(filepath, data_only=True)
        texts = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = []
            for row in ws.iter_rows(values_only=True):
                cells = [str(c) if c is not None else "" for c in row]
                rows.append("\t".join(cells))
            texts.append(f"Sheet: {sheet_name}\n" + "\n".join(rows))
        return {"text": "\n\n".join(texts)}
    except ImportError:
        # Fallback: try CSV-like reading
        return _extract_text(filepath)
    except Exception as e:
        return {"error": True, "message": f"Failed to read Excel file: {str(e)}"}


def _extract_image(filepath: str) -> dict:
    """Extract text from image using Gemini vision."""
    try:
        img = Image.open(filepath)
        client = genai.Client()
        response = client.models.generate_content(
            model=GEMINI_MODEL,
            contents=["Extract ALL text from this image exactly as written. Preserve formatting, numbers, and special characters:", img],
        )
        return {"text": response.text}
    except Exception as e:
        return {"error": True, "message": f"Image extraction failed: {str(e)}"}
