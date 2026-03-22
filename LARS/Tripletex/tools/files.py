import os
import shutil
import logging

from kreuzberg import extract_file_sync

log = logging.getLogger(__name__)

SAVE_DIR = os.path.join(os.path.dirname(__file__), "..", "extracted_files")
os.makedirs(SAVE_DIR, exist_ok=True)


def _save_copy(filepath: str) -> None:
    """Save a copy of the input file to extracted_files/ for debugging."""
    try:
        dest = os.path.join(SAVE_DIR, os.path.basename(filepath))
        shutil.copy2(filepath, dest)
        log.info(f"Saved file copy to {dest}")
    except Exception as e:
        log.warning(f"Failed to save file copy: {e}")


def build_file_tools(files_dir: str) -> dict:
    """Build file extraction tools. files_dir is the per-request temp directory."""

    def extract_file_content(filename: str) -> dict:
        """Extract text content from a PDF or image file attachment.

        Args:
            filename: The name of the attached file to extract text from. Supports PDF and image files (PNG, JPG, JPEG).

        Returns:
            The extracted text content, or an error message.
        """
        filepath = os.path.join(files_dir, filename)
        if not os.path.exists(filepath):
            return {"error": True, "message": f"File not found: {filename}"}

        lower = filename.lower()
        if lower.endswith(".pdf"):
            return _extract_pdf(filepath)
        elif lower.endswith((".png", ".jpg", ".jpeg", ".webp")):
            return _extract_image(filepath)
        elif lower.endswith((".csv", ".txt", ".tsv")):
            return _extract_text(filepath)
        elif lower.endswith((".xlsx", ".xls")):
            return _extract_excel(filepath)
        else:
            # Try reading as plain text
            return _extract_text(filepath)

    return {"extract_file_content": extract_file_content}


def _extract_pdf(filepath: str) -> dict:
    """Extract text from PDF using kreuzberg."""
    _save_copy(filepath)
    try:
        result = extract_file_sync(filepath)
        if result.content.strip():
            log.info(f"Successfully extracted PDF with kreuzberg: {filepath}")
            preview = result.content[:2000]
            log.info(f"PDF extracted text ({len(result.content)} chars):\n{preview}")
            return {"text": result.content}
        else:
            return {"error": True, "message": f"Kreuzberg returned empty text for {filepath}"}
    except Exception as e:
        return {"error": True, "message": f"PDF extraction failed: {str(e)}"}


def _extract_text(filepath: str) -> dict:
    """Extract content from text/CSV files."""
    _save_copy(filepath)
    try:
        for encoding in ("utf-8", "latin-1", "cp1252"):
            try:
                with open(filepath, "r", encoding=encoding) as f:
                    text = f.read()
                preview = text[:2000]
                log.info(f"Text file extracted ({len(text)} chars, {encoding}):\n{preview}")
                return {"text": text}
            except UnicodeDecodeError:
                continue
        return {"error": True, "message": "Could not decode text file"}
    except Exception as e:
        return {"error": True, "message": f"Failed to read text file: {str(e)}"}


def _extract_excel(filepath: str) -> dict:
    """Extract content from Excel files."""
    _save_copy(filepath)
    try:
        import openpyxl
        wb = openpyxl.load_workbook(filepath, data_only=True)
        texts = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = []
            for row in ws.iter_rows(values_only=True):
                cells = [str(c) if c is not None else "" for c in row]
                rows.append("\t".join(cells))
            texts.append(f"Sheet: {sheet_name}\n" + "\n".join(rows))
        full = "\n\n".join(texts)
        preview = full[:2000]
        log.info(f"Excel extracted ({len(full)} chars, {len(wb.sheetnames)} sheets):\n{preview}")
        return {"text": full}
    except ImportError:
        # Fallback: try CSV-like reading
        return _extract_text(filepath)
    except Exception as e:
        return {"error": True, "message": f"Failed to read Excel file: {str(e)}"}


def _extract_image(filepath: str) -> dict:
    """Extract text from image using kreuzberg."""
    _save_copy(filepath)
    try:
        result = extract_file_sync(filepath)
        if result.content.strip():
            preview = result.content[:2000]
            log.info(f"Image extracted text ({len(result.content)} chars):\n{preview}")
            return {"text": result.content}
        else:
            return {"error": True, "message": f"Kreuzberg returned empty text for {filepath}"}
    except Exception as e:
        return {"error": True, "message": f"Image extraction failed: {str(e)}"}
